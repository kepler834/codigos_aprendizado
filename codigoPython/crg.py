import openpyxl

dicionario= {'E' : 10.0, 'B' : 7.5, 'R' : 5.0, 'I' : 2.5, 'S' : 0.0}

notas = openpyxl.load_workbook('D:/UFPA/Diversos/notas.xlsx')
sheet = notas.worksheets[0]

sum_numerador = 0
sum_denominador = 0

for linha in range(1, sheet.max_row + 1):
    for coluna in range(1, sheet.max_column + 1):
        if sheet.cell(row=linha, column=coluna).value == "Conceito":
            i = linha+1
            while sheet.cell(row=i, column=coluna).value in dicionario:
                nota = dicionario[sheet.cell(row=i, column=coluna).value]
                ch = int(str(sheet.cell(row=i, column=coluna).comment)[16:19])
                sum_numerador += nota*ch
                sum_denominador += ch
                i+=1
            sheet.cell(row=i, column=coluna).value = float(sum_numerador/sum_denominador)
notas.save('D:/UFPA/Diversos/notas.xlsx')